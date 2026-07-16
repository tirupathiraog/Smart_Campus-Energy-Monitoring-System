from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer
)

from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.units import inch
from reportlab.lib.colors import HexColor
from datetime import datetime
from reportlab.lib.pagesizes import A4


from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from fastapi.responses import FileResponse
import os
import traceback
import threading
import time
import pandas as pd
from datetime import datetime

from fastapi import Query
from fastapi import FastAPI, Request
from fastapi.responses import HTMLResponse
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
from sqlalchemy import text

from database import engine

app = FastAPI()
from fastapi.middleware.cors import CORSMiddleware

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
templates = Jinja2Templates(directory="templates")


# ---------------- MODEL ----------------
class Energy(BaseModel):
    building: str
    room: str
    timestamp: datetime
    power_kwh: float


# ---------------- UPLOAD ----------------
@app.post("/upload")
def upload_data(data: Energy):

    with engine.begin() as conn:

        # Save energy reading
        conn.execute(text("""
            INSERT INTO energy_readings
            (building, room, timestamp, power_kwh)
            VALUES
            (:building, :room, :timestamp, :power_kwh)
        """), data.dict())

        # -------- Alert Logic --------
        if 50 <= data.power_kwh < 80:
            risk_score = 60
            alert_type = "Moderate Spike"

        elif data.power_kwh >= 80:
            risk_score = 90
            alert_type = "High Usage"

        else:
            return {"message": "stored"}

        # Save alert
        conn.execute(text("""
            INSERT INTO energy_alerts
            (timestamp, building, room, power_kwh, risk_score, alert_type)
            VALUES
            (:timestamp, :building, :room, :power_kwh, :risk_score, :alert_type)
        """), {
            "timestamp": data.timestamp,
            "building": data.building,
            "room": data.room,
            "power_kwh": data.power_kwh,
            "risk_score": risk_score,
            "alert_type": alert_type
        })

    return {"message": "stored"}


# ---------------- BACKGROUND ----------------
def background_detection():
    while True:
        try:
            df = pd.read_sql(
                "SELECT * FROM energy_readings ORDER BY timestamp DESC",
                engine
            )

            if not df.empty:

                df['timestamp'] = pd.to_datetime(df['timestamp'])
                df['avg'] = df['power_kwh'].rolling(window=5).mean()

                latest = df.iloc[0]
                avg_value = df['avg'].iloc[-1]

                # -------- LOGIC --------
                if latest['power_kwh'] < 5:
                    risk_score = 10
                    alert_type = "Low Usage"

                elif latest['power_kwh'] < 50:
                    if pd.notna(avg_value) and latest['power_kwh'] > avg_value + 5:
                        risk_score = 40
                        alert_type = "Moderate Spike"
                    else:
                        risk_score = 20
                        alert_type = "Stable"

                else:
                    risk_score = 80
                    alert_type = "High Usage"

                # ✅ ONLY STORE REAL ALERTS
                if risk_score < 40:
                    continue

                # -------- DUPLICATE CHECK --------
                with engine.begin() as conn:
                    exists = conn.execute(text("""
                        SELECT COUNT(*) FROM energy_alerts
                        WHERE building = :building
                        AND room = :room
                        AND alert_type = :alert_type
                        AND DATE(timestamp) = CURRENT_DATE
                    """), {
                        "building": latest['building'],
                        "room": latest['room'],
                        "alert_type": alert_type
                    }).scalar()

                    if exists == 0:
                        conn.execute(text("""
                            INSERT INTO energy_alerts
                            (timestamp, building, room, power_kwh, risk_score, alert_type)
                            VALUES (:timestamp, :building, :room, :power_kwh, :risk_score, :alert_type)
                        """), {
                            "timestamp": datetime.now(),
                            "building": latest["building"],
                            "room": latest["room"],
                            "power_kwh": latest["power_kwh"],
                            "risk_score": risk_score,
                            "alert_type": alert_type
                        })
        except Exception as e:
            print("========== BACKGROUND ERROR ==========")
            traceback.print_exc()

        time.sleep(10)


#@app.on_event("startup")
#def start_bg():
#    t = threading.Thread(target=background_detection)
#    t.daemon = True
#    t.start()


# ---------------- APIs ----------------

@app.get("/energy/history")
def energy(building: str = Query("All")):

    if building == "All":

        query = """
        SELECT *
        FROM energy_readings
        ORDER BY timestamp DESC
        LIMIT 50
        """

        df = pd.read_sql(query, engine)

    else:

        query = """
        SELECT *
        FROM energy_readings
        WHERE building = %(building)s
        ORDER BY timestamp DESC
        LIMIT 50
        """

        df = pd.read_sql(
            query,
            engine,
            params={"building": building}
        )

    return df.to_dict(orient="records")


@app.get("/alerts")
def alerts():

    df = pd.read_sql("""
        SELECT *
        FROM energy_alerts
        ORDER BY timestamp DESC
    """, engine)

    return df.to_dict(orient="records")


# ✅ FIXED SUMMARY (NO undefined)
@app.get("/summary")
def summary():

    energy_df = pd.read_sql("SELECT * FROM energy_readings", engine)
    alerts_df = pd.read_sql("SELECT * FROM energy_alerts ORDER BY timestamp DESC", engine)

    total_readings = len(energy_df)
    total_alerts = len(alerts_df)

    # ✅ default LOW (fix undefined)
    current_risk = "LOW"

    if not alerts_df.empty:
        latest = alerts_df.iloc[0]["risk_score"]

        if latest >= 70:
            current_risk = "HIGH"
        elif latest >= 30:
            current_risk = "MEDIUM"

    return {
        "total_readings": total_readings,
        "total_alerts": total_alerts,
        "current_risk": current_risk
    }


@app.get("/analytics")
def analytics(building: str = Query("All")):

    where_clause = ""
    params = {}

    if building != "All":
        where_clause = "WHERE building = :building"
        params["building"] = building

    query = f"""
    SELECT

        COALESCE(SUM(
            CASE
                WHEN DATE(timestamp)=CURRENT_DATE
                THEN power_kwh
            END
        ),0) AS today,

        COALESCE(SUM(
            CASE
                WHEN timestamp >= CURRENT_DATE - INTERVAL '7 days'
                THEN power_kwh
            END
        ),0) AS week,

        COALESCE(SUM(
            CASE
                WHEN DATE_TRUNC('month',timestamp)=DATE_TRUNC('month',CURRENT_DATE)
                THEN power_kwh
            END
        ),0) AS month,

        COALESCE(AVG(power_kwh),0) AS average

    FROM energy_readings

    {where_clause}
    """

    df = pd.read_sql(text(query), engine, params=params)

    return {
        "today": round(float(df.iloc[0]["today"]),2),
        "week": round(float(df.iloc[0]["week"]),2),
        "month": round(float(df.iloc[0]["month"]),2),
        "average": round(float(df.iloc[0]["average"]),2)
    }

from fastapi import Query

@app.get("/analytics/buildings")
def building_comparison(building: str = Query("All")):

    where_clause = ""
    params = {}

    if building != "All":
        where_clause = "WHERE building = :building"
        params["building"] = building

    query = f"""
    SELECT
        building,
        ROUND(SUM(power_kwh)::numeric,2) AS total
    FROM energy_readings
    {where_clause}
    GROUP BY building
    ORDER BY total DESC
    """

    df = pd.read_sql(text(query), engine, params=params)

    return df.to_dict(orient="records")

# ✅ FIXED BLOCK DATA (NO undefined)
@app.get("/block/summary")
def block_summary():

    query = """
    SELECT DISTINCT ON (building)
        building,
        room,
        power_kwh
    FROM energy_readings
    ORDER BY building, timestamp DESC
    """

    df = pd.read_sql(query, engine)

    return df.to_dict(orient="records")


@app.get("/active-alerts")
def active_alerts():

    alerts_df = pd.read_sql(
        "SELECT * FROM energy_alerts",
        engine
    )

    active = len(
        alerts_df[
            (alerts_df["alert_type"] == "High Usage") &
            (alerts_df["acknowledged"] == False)
        ]
    )

    return {
        "active_alerts": active
    }

@app.post("/acknowledge/{alert_id}")
def acknowledge_alert(alert_id: int):

    with engine.begin() as conn:

        conn.execute(text("""
            UPDATE energy_alerts
            SET acknowledged = TRUE
            WHERE id = :id
        """), {"id": alert_id})

    return {"message": "Acknowledged"}

@app.get("/download/excel")
def download_excel():

    alerts_df = pd.read_sql(
        "SELECT * FROM energy_alerts ORDER BY timestamp DESC",
        engine
    )

    energy_df = pd.read_sql(
        "SELECT * FROM energy_readings ORDER BY timestamp DESC",
        engine
    )

    summary_df = pd.DataFrame({
        "Metric": [
            "Total Readings",
            "Total Alerts",
            "High Usage Alerts",
            "Moderate Alerts",
            "Acknowledged Alerts",
            "Active Alerts"
        ],
        "Value": [
            len(energy_df),
            len(alerts_df),
            len(alerts_df[alerts_df["alert_type"] == "High Usage"]),
            len(alerts_df[alerts_df["alert_type"] == "Moderate Spike"]),
            len(alerts_df[alerts_df["acknowledged"] == True]),
            len(
                alerts_df[
                    (alerts_df["alert_type"] == "High Usage") &
                    (alerts_df["acknowledged"] == False)
                ]
            )
        ]
    })

    building_df = (
        energy_df
        .groupby("building")["power_kwh"]
        .sum()
        .reset_index()
    )

    filename = "Smart_Campus_Energy_Report.xlsx"

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:

        summary_df.to_excel(
            writer,
            sheet_name="Dashboard Summary",
            index=False
        )

        energy_df.to_excel(
            writer,
            sheet_name="Energy Readings",
            index=False
        )

        alerts_df.to_excel(
            writer,
            sheet_name="Alert History",
            index=False
        )

        building_df.to_excel(
            writer,
            sheet_name="Building Summary",
            index=False
        )

        workbook = writer.book

        header_fill = PatternFill(
            start_color="1F4E78",
            end_color="1F4E78",
            fill_type="solid"
        )

        header_font = Font(
            bold=True,
            color="FFFFFF"
        )

        center = Alignment(
            horizontal="center",
            vertical="center"
        )

        for sheet in workbook.sheetnames:

            ws = workbook[sheet]

            for cell in ws[1]:

                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center

            for column in ws.columns:

                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:

                    cell.alignment = center

                    try:
                        if cell.value is not None:
                            max_length = max(
                                max_length,
                                len(str(cell.value))
                            )
                    except:
                        pass

                ws.column_dimensions[column_letter].width = max_length + 5

    return FileResponse(
        path=filename,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.get("/download/pdf")
def download_pdf():
    return generate_pdf(engine)

@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request):
    return templates.TemplateResponse("dashboard.html", {"request": request})
@app.get("/about", response_class=HTMLResponse)
def about(request: Request):
    return templates.TemplateResponse(
        "about.html",
        {"request": request}
    )